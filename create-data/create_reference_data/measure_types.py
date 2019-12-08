# Import libraries
from openpyxl import Workbook, load_workbook
import common.globals as g
from common.measure_type import measure_type
from common.application import application

# Determine file paths
g.app.d("Creating measure types XML", False)
g.app.d("Reading measure types source", True)
g.app.get_templates()
g.app.get_profile_paths(sub_folder="measure_types", default_filename="measure_types")

# Open the Excel workbook
wb = load_workbook(filename=g.app.input_filepath, read_only=True)

# Updated measure types
ws = wb['Updated']
row_count = ws.max_row
for i in range(2, row_count + 1):
    MEASURE_TYPE_ID = ws.cell(row=i, column=1).value
    DESCRIPTION = ws.cell(row=i, column=2).value
    obj = measure_type(MEASURE_TYPE_ID, DESCRIPTION, "", "", "", "", "", "", "", "", "update")
    g.app.measure_type_list.append(obj)

# New measure types
ws = wb['New']
row_count = ws.max_row
for i in range(2, row_count + 1):
    MEASURE_TYPE_ID = ws.cell(row=i, column=1).value
    DESCRIPTION = ws.cell(row=i, column=2).value
    VALIDITY_START_DATE = ws.cell(row=i, column=3).value
    TRADE_MOVEMENT_CODE = ws.cell(row=i, column=4).value
    PRIORITY_CODE = ws.cell(row=i, column=5).value
    MEASURE_COMPONENT_APPLICABLE_CODE = ws.cell(row=i, column=6).value
    ORIGIN_DEST_CODE = ws.cell(row=i, column=7).value
    ORDER_NUMBER_CAPTURE_CODE = ws.cell(row=i, column=8).value
    MEASURE_EXPLOSION_LEVEL = ws.cell(row=i, column=9).value
    MEASURE_TYPE_SERIES_ID = ws.cell(row=i, column=10).value
    obj = measure_type(MEASURE_TYPE_ID, DESCRIPTION, VALIDITY_START_DATE, TRADE_MOVEMENT_CODE, PRIORITY_CODE, MEASURE_COMPONENT_APPLICABLE_CODE, ORIGIN_DEST_CODE, ORDER_NUMBER_CAPTURE_CODE, MEASURE_EXPLOSION_LEVEL, MEASURE_TYPE_SERIES_ID, "insert")
    g.app.measure_type_list.append(obj)

env = g.app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(g.app.base_envelope_id))
out = ""
for obj in g.app.measure_type_list:
    obj.writeXML(g.app)
    out += obj.xml

# Write the XML
out = env.replace("[BODY]", out)
f = open(g.app.output_filepath, "w", encoding="utf-8")
f.write(out)
f.close()

# Validate the XML
g.app.validate()

# Update the configuration file, as required
g.app.set_config()
