# Import libraries
from openpyxl import Workbook, load_workbook
import common.globals as g
from common.base_regulation import base_regulation
from common.application import application

# Determine file paths
g.app.d("Creating base regulation XML", False)
g.app.d("Reading base regulations source", True)
g.app.get_templates()
g.app.get_profile_paths(sub_folder="base_regulations", default_filename="base_regulations")

# Open the Excel workbook
wb = load_workbook(filename=g.app.input_filepath, read_only=True, data_only=True)

# Read the new items
ws = wb['New']
row_count = ws.max_row
for i in range(2, row_count + 1):
    BASE_REGULATION_ID = ws.cell(row=i, column=1).value
    VALIDITY_START_DATE = ws.cell(row=i, column=2).value
    REGULATION_GROUP_ID = ws.cell(row=i, column=3).value
    LEGISLATION_ID = ws.cell(row=i, column=4).value
    URL = ws.cell(row=i, column=5).value
    INFORMATION_TEXT = ws.cell(row=i, column=6).value
    OMIT = ws.cell(row=i, column=7).value
    if OMIT != "Y":
        obj = base_regulation(BASE_REGULATION_ID, VALIDITY_START_DATE, REGULATION_GROUP_ID, LEGISLATION_ID, URL, INFORMATION_TEXT, "insert")
        g.app.base_regulations_list.append(obj)

# Read the updated items
ws = wb['Updated']
row_count = ws.max_row
for i in range(2, row_count + 1):
    BASE_REGULATION_ID = ws.cell(row=i, column=1).value
    VALIDITY_START_DATE = ws.cell(row=i, column=2).value
    REGULATION_GROUP_ID = ws.cell(row=i, column=3).value
    LEGISLATION_ID = ws.cell(row=i, column=4).value
    URL = ws.cell(row=i, column=5).value
    INFORMATION_TEXT = ws.cell(row=i, column=6).value
    obj = base_regulation(BASE_REGULATION_ID, VALIDITY_START_DATE, REGULATION_GROUP_ID, LEGISLATION_ID, URL, INFORMATION_TEXT, "update")
    g.app.base_regulations_list.append(obj)

# Write the XML
g.app.d("Writing base regulations XML", True)
env = g.app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(g.app.base_envelope_id))
out = ""
for obj in g.app.base_regulations_list:
    obj.writeXML(g.app)
    out += obj.xml

out = env.replace("[BODY]", out)
f = open(g.app.output_filepath, "w", encoding="utf-8")
f.write(out)
f.close()

# Validate the XML
g.app.validate()

# Update the configuration file, as required
g.app.set_config()
