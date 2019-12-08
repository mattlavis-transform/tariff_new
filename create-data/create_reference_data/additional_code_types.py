# Import libraries
from openpyxl import Workbook, load_workbook
import common.globals as g
from common.additional_code_type import additional_code_type
from common.application import application

# Determine file paths
g.app.d("Creating additional codes XML", False)
g.app.d("Reading additional codes source", True)
g.app.get_templates()
g.app.get_profile_paths(sub_folder="additional_code_types", default_filename="additional_code_types")

# Open the Excel workbook
wb = load_workbook(filename=g.app.input_filepath, read_only=True)

# Read the updated items
ws = wb['Updated']
row_count = ws.max_row
for i in range(2, row_count + 1):
    ADDITIONAL_CODE_TYPE_ID = ws.cell(row=i, column=1).value
    DESCRIPTION = ws.cell(row=i, column=2).value
    APPLICATION_CODE = "n/a"
    act = additional_code_type(ADDITIONAL_CODE_TYPE_ID, DESCRIPTION, APPLICATION_CODE, "update")
    g.app.additional_code_type_list.append(act)

# Read the new items
ws = wb['New']
row_count = ws.max_row
for i in range(2, row_count + 1):
    ADDITIONAL_CODE_TYPE_ID = ws.cell(row=i, column=1).value
    DESCRIPTION = ws.cell(row=i, column=2).value
    APPLICATION_CODE = ws.cell(row=i, column=3).value
    f = additional_code_type(ADDITIONAL_CODE_TYPE_ID, DESCRIPTION, APPLICATION_CODE, "insert")
    g.app.additional_code_type_list.append(f)

# Write the XML
g.app.d("Writing additional codes XML", True)
env = g.app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(g.app.base_envelope_id))
out = ""
for obj in g.app.additional_code_type_list:
    obj.writeXML(g.app)
    out += obj.xml

out = env.replace("[BODY]", out)
f = open(g.app.output_filepath, "w", encoding="utf-8")
f.write(out)
f.close()

# Validate the XML
g.app.validate()

# Update the configuration file, as required
g.app.set_config()
