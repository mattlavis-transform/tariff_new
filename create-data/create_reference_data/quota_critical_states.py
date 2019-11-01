import psycopg2
import sys
import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.objects as o
from common.quota_definition import quota_definition
from common.application import application

try:
	profile = sys.argv[1]
except:
	profile = "quota_critical_states"

app = o.app
app.d("Writing critical states XML", False)
app.get_templates()

fname = os.path.join(app.SOURCE_DIR, profile + ".xlsx")
app.d("Reading Excel source file")
wb = load_workbook(filename = fname, read_only = True, data_only = True)

ws = wb['critical states']
row_count = ws.max_row

quota_definition_list = []
for i in range(2, row_count + 1):
	quota_order_number_id   = ws.cell(row = i, column = 1).value
	critical_state          = ws.cell(row = i, column = 2).value

	# Get the latest quota order number origin SID
	sql = """select quota_definition_sid, quota_order_number_id, validity_start_date, validity_end_date,
	quota_order_number_sid, volume, initial_volume, measurement_unit_code, maximum_precision,
	critical_state, critical_threshold, monetary_unit_code, measurement_unit_qualifier_code, description
	from quota_definitions where quota_order_number_id = %s and validity_start_date >= %s;"""

	params = []
	params.append(quota_order_number_id)
	params.append(o.app.critical_date_plus_one_string)
	cur = o.app.conn.cursor()
	cur.execute(sql, params)
	rows = cur.fetchall()
	quota_order_number_origin_sid = None

	for row in rows:
		qd = quota_definition()

		qd.quota_definition_sid            = row[0]
		qd.quota_order_number_id           = row[1]
		qd.validity_start_date             = row[2]
		qd.validity_end_date               = row[3]
		qd.quota_order_number_sid          = row[4]
		qd.volume                          = row[5]
		qd.initial_volume                  = row[6]
		qd.measurement_unit_code           = row[7]
		qd.maximum_precision               = row[8]
		qd.critical_state                  = critical_state
		qd.critical_threshold              = row[10]
		qd.monetary_unit_code              = row[11]
		qd.measurement_unit_qualifier_code = row[12]
		qd.description                     = row[13]

		quota_definition_list.append(qd)




filename = os.path.join(app.XML_DIR, profile + ".xml")
app.d("Writing XML file to " + filename)
env = app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(app.base_envelope_id))
out = ""
for obj in quota_definition_list:
	out += obj.xml()

out = env.replace("[BODY]", out)
f = open(filename, "w", encoding="utf-8")
f.write(out)
f.close()

app.validate(filename)
app.set_config()
