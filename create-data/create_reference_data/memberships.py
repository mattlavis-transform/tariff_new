# Import libraries
from openpyxl import Workbook, load_workbook
import common.globals as g
from common.membership import membership
from common.application import application

# Determine file paths
g.app.d("Creating memberships XML", False)
g.app.d("Reading memberships source", True)
g.app.get_templates()
g.app.get_profile_paths(sub_folder="memberships", default_filename="memberships")

# Open the Excel workbook
wb = load_workbook(filename=g.app.input_filepath, read_only=True)

# All memberships (commencements and terminations) in one tab
ws = wb['Memberships']
row_count = ws.max_row

for i in range(2, row_count + 1):
    PARENT_SID = ws.cell(row=i, column=1).value
    PARENT_ID = ws.cell(row=i, column=2).value
    CHILD_SID = ws.cell(row=i, column=3).value
    CHILD_ID = ws.cell(row=i, column=4).value
    START_DATE = ws.cell(row=i, column=5).value
    END_DATE = ws.cell(row=i, column=6).value
    UPDATE_TYPE = ws.cell(row=i, column=7).value

    obj = membership(PARENT_SID, PARENT_ID, CHILD_SID, CHILD_ID, START_DATE, END_DATE, UPDATE_TYPE)
    g.app.membership_list.append(obj)

env = g.app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(g.app.base_envelope_id))
out = ""
for obj in g.app.membership_list:
    obj.writeXML(g.app)
    out += obj.xml

out = env.replace("[BODY]", out)
f = open(g.app.output_filepath, "w", encoding="utf-8")
f.write(out)
f.close()

# Validate the XML
g.app.validate()

# Update the configuration file, as required
g.app.set_config()
