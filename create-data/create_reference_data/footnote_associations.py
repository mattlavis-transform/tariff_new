import psycopg2
import sys
import os
import xmlschema
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.globals as g
from common.footnote_association import footnote_association
from common.application import application

try:
    profile = sys.argv[1]
except:
    profile = "footnote_associations"

app = g.app
app.get_templates()
app.get_profile(profile)

path = os.path.join(app.SOURCE_DIR, "footnote_associations")
fname = os.path.join(path, profile + ".xlsx")
wb = load_workbook(filename=fname, read_only=True)
ws = wb['measure associations']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
    footnote_type_id = ws.cell(row=i, column=1).value
    footnote_id = ws.cell(row=i, column=2).value
    measure_type_id = ws.cell(row=i, column=3).value
    geographical_area_id = ws.cell(row=i, column=4).value
    goods_nomenclature_item_id = ws.cell(row=i, column=5).value

    f = footnote_association(footnote_type_id, footnote_id, measure_type_id,
                             geographical_area_id, goods_nomenclature_item_id)
    app.footnote_associations_list.append(f)


env = app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(app.base_envelope_id))
out = ""
for f in app.footnote_associations_list:
    f.writeXML(app)
    out += f.xml

out = env.replace("[BODY]", out)
filename = os.path.join(app.XML_DIR, profile + ".xml")
f = open(filename, "w", encoding="utf-8")
f.write(out)
f.close()

schema_path = os.path.join(app.SCHEMA_DIR, "envelope.xsd")
my_schema = xmlschema.XMLSchema(schema_path)
try:
    if my_schema.is_valid(filename):
        print("The file validated successfully.")
    else:
        print("The file did not validate.")
except:
    print("The file did not validate and crashed the validator.")
app.set_config()
