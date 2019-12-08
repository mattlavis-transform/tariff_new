import psycopg2
import sys
import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle


import common.globals as g
from common.quota_order_number import quota_order_number
from common.application import application

try:
    profile = sys.argv[1]
except:
    profile = "quota_objects"


app = g.app
app.d("Writing quota object XML", False)
app.get_templates()
app.get_regulation_profile(profile)


fname = os.path.join(app.SOURCE_DIR, profile + ".xlsx")
app.d("Reading Excel source file")
wb = load_workbook(filename=g.app.excel_source, read_only=True, data_only=True)

ws = wb['objects']
row_count = ws.max_row
col_count = ws.max_column

out = ""

for i in range(2, row_count + 1):
    item = ws.cell(row=i, column=1).value
    instance = ws.cell(row=i, column=2).value
    action = ws.cell(row=i, column=3).value
    change = ws.cell(row=i, column=4).value

    if item == "Quota order number":
        if action == "Terminate":
            q = quota_order_number()
            q.quota_order_number_id = instance
            q.get_data_from_id()
            q.terminate(change)
            out += q.xml()

        elif action == "Start":
            q = quota_order_number()
            q.quota_order_number_id = instance
            q.get_next_sid()
            q.start(change)
            out += q.xml()

filename = os.path.join(app.XML_DIR, profile + ".xml")
app.d("Writing XML file to " + filename)
env = app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(app.base_envelope_id))
out = env.replace("[BODY]", out)
f = open(filename, "w", encoding="utf-8")
f.write(out)
f.close()

app.validate(filename)
app.set_config()
