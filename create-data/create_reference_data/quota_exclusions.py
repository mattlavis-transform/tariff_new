import psycopg2
import sys
import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.globals as g
from common.quota_order_number_origin_exclusions import quota_order_number_origin_exclusions
from common.application import application

try:
    profile = sys.argv[1]
except:
    profile = "quota_exclusions"

app = g.app
app.d("Writing regulation XML", False)
app.get_templates()

fname = os.path.join(app.SOURCE_DIR, profile + ".xlsx")
app.d("Reading Excel source file")
wb = load_workbook(filename=fname, read_only=True, data_only=True)

ws = wb['exclusions']
row_count = ws.max_row

quota_order_number_origin_exclusions_list = []
for i in range(2, row_count + 1):
    quota_order_number_id = ws.cell(row=i, column=1).value
    quota_exclusions = ws.cell(row=i, column=2).value
    quota_exclusions_list = quota_exclusions.split(",")

    # Get the latest quota order number origin SID
    sql = """select quota_order_number_origin_sid
    from quota_order_number_origins qono, quota_order_numbers qon
    where qono.quota_order_number_sid = qon.quota_order_number_sid
    and qon.quota_order_number_id = %s order by qono.validity_start_date desc
    limit 1;"""

    params = []
    params.append(quota_order_number_id)
    cur = g.app.conn.cursor()
    cur.execute(sql, params)
    rows = cur.fetchall()
    quota_order_number_origin_sid = None
    if len(rows) > 0:
        row = rows[0]
        quota_order_number_origin_sid = row[0]

    if quota_order_number_origin_sid is None:
        print("No origin found")
        sys.exit()

    # Check the exclusion does not already exist: cannot recreate
    sql = """select ga.geographical_area_id
    from quota_order_number_origin_exclusions qonoe, geographical_areas ga
    where qonoe.excluded_geographical_area_sid = ga.geographical_area_sid
    and quota_order_number_origin_sid = %s order by 1;"""
    params = []
    params.append(str(quota_order_number_origin_sid))
    cur = g.app.conn.cursor()
    cur.execute(sql, params)
    rows = cur.fetchall()
    for row in rows:
        geographical_area_id = row[0]
        for new_exclusion in quota_exclusions_list:
            if new_exclusion == geographical_area_id:
                print("Exclusion", geographical_area_id, "already exists on quota", quota_order_number_id)

    for new_excluded_geographical_id in quota_exclusions_list:
        ex = quota_order_number_origin_exclusions(new_excluded_geographical_id, quota_order_number_origin_sid)
        ex.quota_order_number_id = quota_order_number_id
        ex.get_measures()
        quota_order_number_origin_exclusions_list.append(ex)

filename = os.path.join(app.XML_DIR, profile + ".xml")
app.d("Writing XML file to " + filename)
env = app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(app.base_envelope_id))
out = ""
for obj in quota_order_number_origin_exclusions_list:
    out += obj.xml()
    out += obj.measure_xml()

out = env.replace("[BODY]", out)
f = open(filename, "w", encoding="utf-8")
f.write(out)
f.close()

app.validate(filename)
app.set_config()
app.set_config2("measure.conditions", app.last_measure_condition_sid)
