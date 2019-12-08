import psycopg2
import sys
import os
import xmlschema
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.globals as g
from common.goods_nomenclature import goods_nomenclature
from common.application import application

app = g.app
app.get_templates()

# Get arguments
if len(sys.argv) > 0:
    fname = sys.argv[1]
    if fname.find(".xlsx") == -1:
        fname += ".xlsx"
fname_out = fname.replace(".xlsx", ".xml")

path = os.path.join(app.SOURCE_DIR, "commodities")
fname = os.path.join(path, fname)
wb = load_workbook(filename=fname, read_only=True)
ws = wb['Updated']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
    goods_nomenclature_item_id = ws.cell(row=i, column=1).value
    productline_suffix = ws.cell(row=i, column=2).value
    description = ws.cell(row=i, column=3).value
    goods_nomenclature_sid = ws.cell(row=i, column=4).value
    change_date = ws.cell(row=i, column=5).value

    if change_date == "":
        change_date = g.app.critical_date_plus_one_string

    obj = goods_nomenclature(goods_nomenclature_sid, goods_nomenclature_item_id, productline_suffix, description, force=True)
    app.goods_nomenclature_list.append(obj)

env = app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(app.base_envelope_id))
out = ""
for obj in app.goods_nomenclature_list:
    obj.writeXML()
    out += obj.xml

out = env.replace("[BODY]", out)
filename = os.path.join(app.XML_DIR, fname_out)
f = open(filename, "w", encoding="utf-8")
f.write(out)
f.close()

schema_path = os.path.join(app.SCHEMA_DIR, "envelope.xsd")
my_schema = xmlschema.XMLSchema(schema_path)
try:
    if my_schema.is_valid(filename):
        print("The file validated successfully.")
    else:
        print("The file did not validate.")
except:
    print("The file did not validate and crashed the validator.")

app.set_config()
app.set_config2("goods.nomenclature.description.periods", app.last_goods_nomenclature_description_period_sid)
