import psycopg2
import sys
import os
import xmlschema
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.globals as g
from common.quota_order_number_origin import quota_order_number_origin
from common.application import application

app = g.app
app.get_templates()

try:
    profile = sys.argv[1]
except:
    profile = "quota_order_number_origins"


fname = os.path.join(app.SOURCE_DIR, profile + ".xlsx")
wb = load_workbook(filename=fname, read_only=True)
ws = wb['origins']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
    quota_order_number_id = ws.cell(row=i, column=1).value
    geographical_area_id = ws.cell(row=i, column=2).value
    validity_start_date = ws.cell(row=i, column=3).value

    obj = quota_order_number_origin(quota_order_number_id, geographical_area_id, validity_start_date)
    app.quota_order_number_origin_list.append(obj)

env = app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(app.base_envelope_id))
out = ""
for obj in app.quota_order_number_origin_list:
    if obj.quota_order_number_id[0:3] != "094":
        obj.writeXML(app)
        out += obj.xml

out = env.replace("[BODY]", out)
filename = os.path.join(app.XML_DIR, profile + ".xml")
f = open(filename, "w", encoding="utf-8")
f.write(out)
f.close()

schema_path = os.path.join(app.SCHEMA_DIR, "envelope.xsd")
my_schema = xmlschema.XMLSchema(schema_path)

try:
    if my_schema.is_valid(filename):
        success = True
        print("The file validated successfully.")
    else:
        success = False
        print("The file did not validate.")
except:
    success = False
    print("The file did not validate and crashed the validator.")


if not(success):
    my_schema.validate(filename)

app.set_config()
