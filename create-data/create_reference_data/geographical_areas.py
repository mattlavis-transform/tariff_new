# Import libraries
from openpyxl import Workbook, load_workbook
import common.globals as g
from common.geographical_area import geographical_area
from common.application import application


# Determine file paths
g.app.d("Creating geographical areas XML", False)
g.app.d("Reading geographical areas source", True)
g.app.get_templates()
g.app.get_profile_paths(sub_folder="geographical_areas", default_filename="geographical_areas")

# Open the Excel workbook
wb = load_workbook(filename=g.app.input_filepath, read_only=True)

# Updated geographical areas
ws = wb['Updated']
row_count = ws.max_row
for i in range(2, row_count + 1):
    GEOGRAPHICAL_AREA_SID = ws.cell(row=i, column=1).value
    GEOGRAPHICAL_AREA_ID = ws.cell(row=i, column=2).value
    DESCRIPTION = ws.cell(row=i, column=3).value
    obj = geographical_area(GEOGRAPHICAL_AREA_SID, GEOGRAPHICAL_AREA_ID, "", DESCRIPTION, "update")
    g.app.geographical_area_list.append(obj)

# New geographical areas
ws = wb['New']
row_count = ws.max_row
for i in range(2, row_count + 1):
    GEOGRAPHICAL_AREA_SID = ws.cell(row=i, column=1).value
    GEOGRAPHICAL_AREA_ID = ws.cell(row=i, column=2).value
    DESCRIPTION = ws.cell(row=i, column=3).value
    GEOGRAPHICAL_AREA_CODE = ws.cell(row=i, column=4).value

    """ AREA CODES
    0 Country
    1 Geographical area group
    2 Region
    """

    obj = geographical_area(GEOGRAPHICAL_AREA_SID, GEOGRAPHICAL_AREA_ID, GEOGRAPHICAL_AREA_CODE, DESCRIPTION, "insert")
    g.app.geographical_area_list.append(obj)

env = g.app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(g.app.base_envelope_id))
out = ""
for obj in g.app.geographical_area_list:
    obj.writeXML(g.app)
    out += obj.xml

# Write the XML
out = env.replace("[BODY]", out)
f = open(g.app.output_filepath, "w", encoding="utf-8")
f.write(out)
f.close()

# Validate the XML
g.app.validate()

# Update the configuration file, as required
g.app.set_config()
