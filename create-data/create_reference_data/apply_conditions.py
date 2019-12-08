import psycopg2
import sys
import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.globals as g
from common.measure_condition import measure_condition
from common.application import application

try:
    profile = sys.argv[1]
except:
    profile = "measure_conditions_reliefs"


app = g.app
app.d("Writing regulation XML", False)
app.get_templates()
app.get_regulation_profile(profile)


fname = os.path.join(app.SOURCE_DIR, profile + ".xlsx")
app.d("Reading Excel source file")
wb = load_workbook(filename=g.app.excel_source, read_only=True, data_only=True)

ws = wb['New']
row_count = ws.max_row
col_count = ws.max_column
filename = os.path.join(app.XML_DIR, app.xml_file)

app.object_list = []
for i in range(2, row_count + 1):
    MEASURE_TYPE_ID = ws.cell(row=i, column=1).value
    UPDATE_TYPE = ws.cell(row=i, column=2).value
    CONDITION_CODE = ws.cell(row=i, column=3).value
    COMPONENT_SEQUENCE_NUMBER = ws.cell(row=i, column=4).value
    ACTION_CODE = ws.cell(row=i, column=5).value
    CERTIFICATE_TYPE_CODE = ws.cell(row=i, column=6).value
    CERTIFICATE_CODE = ws.cell(row=i, column=7).value

    sql = "select measure_sid from measures where measure_type_id = %s"
    params = []
    params.append(str(MEASURE_TYPE_ID))

    cur = g.app.conn.cursor()
    cur.execute(sql, params)
    rows = cur.fetchall()
    if len(rows) > 0:
        for row in rows:
            measure_sid = str(row[0])
            obj = measure_condition(measure_sid, UPDATE_TYPE, CONDITION_CODE, COMPONENT_SEQUENCE_NUMBER, ACTION_CODE, CERTIFICATE_TYPE_CODE, CERTIFICATE_CODE)
            app.object_list.append(obj)

app.d("Writing XML file to " + filename)
env = app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(app.base_envelope_id))
out = ""
for obj in app.object_list:
    obj.writeXML()
    out += obj.xml

out = env.replace("[BODY]", out)
f = open(filename, "w", encoding="utf-8")
f.write(out)
f.close()

app.validate(filename)
app.set_config()
app.set_config2("measure.conditions", app.last_measure_condition_sid)
