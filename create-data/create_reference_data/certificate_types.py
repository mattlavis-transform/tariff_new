# Import libraries
from openpyxl import Workbook, load_workbook
import common.globals as g
from common.certificate_type import certificate_type
from common.application import application

# Determine file paths
g.app.d("Creating certificate types XML", False)
g.app.d("Reading certificate types source", True)
g.app.get_templates()
g.app.get_profile_paths(sub_folder="certificate_types", default_filename="certificate_types")

# Open the Excel workbook
wb = load_workbook(filename=g.app.input_filepath, read_only=True)

ws = wb['New']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
    CERTIFICATE_TYPE_ID = ws.cell(row=i, column=1).value
    DESCRIPTION = ws.cell(row=i, column=2).value
    VALIDITY_START_DATE = ws.cell(row=i, column=3).value

    f = certificate_type(CERTIFICATE_TYPE_ID, DESCRIPTION, VALIDITY_START_DATE, "insert")
    g.app.certificate_type_list.append(f)

env = g.app.envelope_XML
env = env.replace("[ENVELOPE_ID]", str(g.app.base_envelope_id))
out = ""
for obj in g.app.certificate_type_list:
    obj.writeXML(g.app)
    out += obj.xml

# Write the XML
out = env.replace("[BODY]", out)
f = open(g.app.output_filepath, "w", encoding="utf-8")
f.write(out)
f.close()


# Validate the XML
g.app.validate()

# Update the configuration file, as required
g.app.set_config()
